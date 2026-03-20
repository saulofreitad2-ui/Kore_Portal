import streamlit as st
import pandas as pd
import json
import openpyxl
from io import BytesIO

# ═══════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Portal do Representante",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ═══════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&display=swap');

/* ── BASE ─────────────────────────────────────────── */
html, body, [class*="css"], * {
    font-family: 'Syne', sans-serif !important;
    color: white;
}
.stApp {
    background: linear-gradient(155deg, #143f50 0%, #1a5268 100%) !important;
    min-height: 100vh;
}
.block-container {
    padding-top: 1.5rem !important;
    max-width: 1400px !important;
}
footer, #MainMenu, header { visibility: hidden; }

/* ── MÉTRICAS ─────────────────────────────────────── */
[data-testid="metric-container"] {
    background: #1e6070 !important;
    border: 1px solid rgba(255,255,255,0.12) !important;
    border-radius: 18px !important;
    padding: 18px 20px !important;
}
[data-testid="metric-container"] > div { gap: 4px !important; }
[data-testid="stMetricLabel"] > div,
[data-testid="stMetricLabel"] p {
    color: rgba(255,255,255,0.6) !important;
    font-size: 11px !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.09em !important;
}
[data-testid="stMetricValue"],
[data-testid="stMetricValue"] > div {
    color: #F47920 !important;
    font-size: 22px !important;
    font-weight: 800 !important;
    letter-spacing: -0.02em !important;
}
[data-testid="stMetricDelta"] svg { display: none; }
[data-testid="stMetricDelta"] > div {
    color: rgba(255,255,255,0.5) !important;
    font-size: 12px !important;
}

/* ── TABS ─────────────────────────────────────────── */
[data-baseweb="tab-list"] {
    background: transparent !important;
    gap: 6px !important;
    border-bottom: 1px solid rgba(255,255,255,0.08) !important;
    padding-bottom: 4px !important;
}
[data-baseweb="tab"] {
    background: rgba(0,0,0,0.25) !important;
    border-radius: 10px !important;
    color: rgba(255,255,255,0.7) !important;
    font-weight: 700 !important;
    font-size: 13px !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
    padding: 8px 18px !important;
    margin: 0 !important;
}
[data-baseweb="tab"]:hover {
    background: rgba(244,121,32,0.15) !important;
    color: white !important;
}
[data-baseweb="tab"][aria-selected="true"] {
    background: #F47920 !important;
    color: white !important;
    border-color: #F47920 !important;
    box-shadow: 0 4px 16px rgba(244,121,32,0.35) !important;
}
[data-baseweb="tab-highlight"],
[data-baseweb="tab-border"] { display: none !important; }

/* ── DATAFRAME ────────────────────────────────────── */
[data-testid="stDataFrame"] > div {
    background: #1a5060 !important;
    border-radius: 14px !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    overflow: hidden !important;
}
[data-testid="stDataFrame"] iframe {
    border-radius: 12px !important;
}

/* ── INPUTS ───────────────────────────────────────── */
[data-testid="stTextInput"] input,
[data-baseweb="input"] input,
input[type="text"],
input[type="password"] {
    background: rgba(0,0,0,0.3) !important;
    border: 1px solid rgba(255,255,255,0.15) !important;
    border-radius: 10px !important;
    color: white !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 14px !important;
    padding: 10px 14px !important;
}
[data-testid="stTextInput"] input:focus {
    border-color: rgba(244,121,32,0.6) !important;
    box-shadow: 0 0 0 2px rgba(244,121,32,0.15) !important;
}
[data-testid="stTextInput"] label,
[data-testid="stTextInput"] p {
    color: rgba(255,255,255,0.7) !important;
    font-size: 12px !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
}
/* Placeholder */
[data-testid="stTextInput"] input::placeholder { color: rgba(255,255,255,0.35) !important; }

/* ── RADIO ────────────────────────────────────────── */
[data-testid="stRadio"] label {
    color: white !important;
    font-size: 13px !important;
    font-weight: 600 !important;
}
[data-testid="stRadio"] > div {
    background: rgba(0,0,0,0.2) !important;
    border-radius: 10px !important;
    padding: 6px 12px !important;
    gap: 16px !important;
}
[data-testid="stRadio"] [data-testid="stMarkdown"] p { color: white !important; }

/* ── SELECT / DROPDOWN ────────────────────────────── */
[data-testid="stSelectbox"] label,
[data-testid="stSelectbox"] p {
    color: rgba(255,255,255,0.7) !important;
    font-size: 12px !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
}
[data-baseweb="select"] > div:first-child {
    background: rgba(0,0,0,0.3) !important;
    border: 1px solid rgba(255,255,255,0.15) !important;
    border-radius: 10px !important;
    color: white !important;
}
[data-baseweb="select"] span { color: white !important; }
[data-baseweb="popover"] { background: #1a5060 !important; }
[data-baseweb="menu"] { background: #1a5060 !important; border: 1px solid rgba(255,255,255,0.1) !important; border-radius: 10px !important; }
[data-baseweb="menu"] li { color: white !important; }
[data-baseweb="menu"] li:hover { background: rgba(244,121,32,0.2) !important; }

/* ── BOTÕES ───────────────────────────────────────── */
.stButton > button {
    background: #F47920 !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 13px !important;
    font-family: 'Syne', sans-serif !important;
    padding: 8px 20px !important;
    transition: opacity 0.15s !important;
}
.stButton > button:hover { opacity: 0.85 !important; }
.stButton > button[kind="secondary"] {
    background: rgba(255,255,255,0.1) !important;
}

/* ── FORM ─────────────────────────────────────────── */
[data-testid="stForm"] {
    background: #1e6070 !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    border-radius: 18px !important;
    padding: 24px !important;
}
[data-testid="stFormSubmitButton"] > button {
    background: #8DC63F !important;
    width: 100% !important;
}

/* ── EXPANDER ─────────────────────────────────────── */
[data-testid="stExpander"] {
    background: #1e6070 !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    border-radius: 14px !important;
}
[data-testid="stExpander"] summary {
    color: white !important;
    font-weight: 700 !important;
}
[data-testid="stExpander"] svg { fill: white !important; }

/* ── FILE UPLOADER ────────────────────────────────── */
[data-testid="stFileUploader"] {
    background: rgba(0,0,0,0.2) !important;
    border: 2px dashed rgba(255,255,255,0.2) !important;
    border-radius: 14px !important;
    padding: 20px !important;
}
[data-testid="stFileUploader"] label { color: white !important; }

/* ── ALERTS ───────────────────────────────────────── */
[data-testid="stAlert"] {
    border-radius: 12px !important;
}

/* ── SCROLLBAR ────────────────────────────────────── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: rgba(255,255,255,0.2); border-radius: 4px; }

/* ── HR ───────────────────────────────────────────── */
hr { border-color: rgba(255,255,255,0.08) !important; }

/* ── MARKDOWN TEXT ────────────────────────────────── */
.stMarkdown p, .stMarkdown span, .stMarkdown div { color: white !important; }

/* ── BADGES E CARDS CUSTOM ────────────────────────── */
.badge {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.04em;
    text-transform: uppercase;
}
.prog-bar-bg {
    background: rgba(0,0,0,0.3);
    border-radius: 99px;
    height: 14px;
    overflow: hidden;
    margin: 8px 0;
}
.prog-bar-fill {
    height: 100%;
    border-radius: 99px;
    background: linear-gradient(90deg, #8DC63F, #a8e05f);
}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# DADOS
# ═══════════════════════════════════════════════════════════════

def safe_float(v):
    try: return float(v or 0)
    except: return 0.0

def fmt_date(v):
    if hasattr(v, 'strftime'): return v.strftime("%d/%m/%Y")
    return str(v or "")

def load_from_excel(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes))

    # Estoque
    ws = wb['Estoque']
    estoque = []
    for r in range(4, ws.max_row + 1):
        m2    = safe_float(ws.cell(r, 8).value)
        rolos = safe_float(ws.cell(r, 9).value)
        if m2 <= 0 and rolos <= 0: continue
        estoque.append({
            "filial": str(ws.cell(r, 1).value or "").strip(),
            "familia": str(ws.cell(r, 2).value or "").strip(),
            "desc_produto": str(ws.cell(r, 3).value or "").strip(),
            "sku": str(ws.cell(r, 4).value or "").strip(),
            "produto": str(ws.cell(r, 5).value or "").strip(),
            "largura": safe_float(ws.cell(r, 6).value),
            "comprimento": safe_float(ws.cell(r, 7).value),
            "m2_disponivel": round(m2, 2),
            "rolos_disponivel": round(rolos, 2),
        })

    # Pedidos
    ws2 = wb['Pedidos']
    pedidos = []
    for r in range(4, ws2.max_row + 1):
        vend = str(ws2.cell(r, 12).value or "").strip()
        if vend in ('', 'Vendedor'): continue
        pedidos.append({
            "filial": str(ws2.cell(r, 1).value or "").strip(),
            "pedido": str(ws2.cell(r, 2).value or "").strip(),
            "uf": str(ws2.cell(r, 3).value or "").strip(),
            "emissao": fmt_date(ws2.cell(r, 4).value),
            "entrega": fmt_date(ws2.cell(r, 5).value),
            "cliente": str(ws2.cell(r, 7).value or "").strip()[:55],
            "segmento": str(ws2.cell(r, 8).value or "").strip(),
            "sku": str(ws2.cell(r, 10).value or "").strip(),
            "produto": str(ws2.cell(r, 11).value or "").strip()[:55],
            "vendedor": vend,
            "status": str(ws2.cell(r, 14).value or "").strip(),
            "valor_vendido": round(safe_float(ws2.cell(r, 15).value), 2),
            "valor_atendido": round(safe_float(ws2.cell(r, 16).value), 2),
            "valor_a_entregar": round(safe_float(ws2.cell(r, 17).value), 2),
            "m2_vendido": round(safe_float(ws2.cell(r, 29).value), 2),
            "m2_atendido": round(safe_float(ws2.cell(r, 30).value), 2),
            "m2_saldo": round(safe_float(ws2.cell(r, 31).value), 2),
        })

    return estoque, pedidos


@st.cache_data
def load_seed_data():
    with open("estoque.json", encoding="utf-8") as f: estoque = json.load(f)
    with open("pedidos.json", encoding="utf-8") as f: pedidos = json.load(f)
    with open("users.json",   encoding="utf-8") as f: users   = json.load(f)
    return estoque, pedidos, users


def init_state():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
        st.session_state.current_user = None

    if "data_loaded" not in st.session_state:
        estoque, pedidos, users = load_seed_data()
        st.session_state.estoque = estoque
        st.session_state.pedidos = pedidos
        st.session_state.users   = users
        st.session_state.data_loaded = True


# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

BRL = lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
M2  = lambda v: f"{v:,.2f} m²".replace(",", "X").replace(".", ",").replace("X", ".")

STATUS_COLOR = {
    "Aguardando Faturamento":           "#f5c842",
    "Aguardando Faturamento - Parcial": "#f5c842",
    "Pedido de Venda em Aberto":        "#4db8d4",
    "Aguardando Separação":             "#F47920",
    "Financeiro Rejeitado":             "#e05555",
    "Aguardando Liberação Estoque":     "#b07aff",
    "Aguardando Liberação Comercial":   "#b07aff",
    "Sem Status":                       "rgba(255,255,255,0.4)",
}

def status_badge(s):
    color = STATUS_COLOR.get(s, "rgba(255,255,255,0.3)")
    return f'<span class="badge" style="background:{color}22;color:{color};border:1px solid {color}55">{s}</span>'

def section_header(title, sub=None):
    st.markdown(f"""
    <div style="margin-bottom:20px">
        <div style="font-weight:800;font-size:22px;letter-spacing:-0.02em;color:white">{title}</div>
        {f'<div style="color:rgba(255,255,255,0.6);font-size:13px;margin-top:4px">{sub}</div>' if sub else ''}
    </div>
    """, unsafe_allow_html=True)

def prog_bar(pct, label_left, label_right):
    color = "#8DC63F" if pct >= 70 else "#f5c842" if pct >= 40 else "#e05555"
    st.markdown(f"""
    <div style="margin:12px 0">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
            <span style="font-weight:700;font-size:14px;color:white">% Atendido da Carteira</span>
            <span style="font-weight:800;font-size:22px;color:{color}">{pct:.1f}%</span>
        </div>
        <div class="prog-bar-bg">
            <div class="prog-bar-fill" style="width:{pct}%;background:linear-gradient(90deg,{color},{color}bb)"></div>
        </div>
        <div style="display:flex;justify-content:space-between;font-size:12px;color:rgba(255,255,255,0.5);margin-top:6px">
            <span>Atendido: {label_left}</span><span>Total: {label_right}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

def card_wrap(content_fn):
    st.markdown('<div style="background:#1e6070;border-radius:18px;border:1px solid rgba(255,255,255,0.1);padding:24px">', unsafe_allow_html=True)
    content_fn()
    st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# LOGIN
# ═══════════════════════════════════════════════════════════════

def render_login():
    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown("""
        <div style="text-align:center;margin:40px 0 32px">
            <div style="width:68px;height:68px;border-radius:20px;background:linear-gradient(135deg,#F47920,#ff9a40);
                display:flex;align-items:center;justify-content:center;margin:0 auto 16px;
                font-size:32px;box-shadow:0 12px 40px rgba(244,121,32,0.35)">📦</div>
            <div style="font-size:28px;font-weight:800;letter-spacing:-0.03em;color:white">Portal do Representante</div>
            <div style="color:rgba(255,255,255,0.6);font-size:14px;margin-top:6px">Gestão de Estoque &amp; Carteira</div>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            login    = st.text_input("Usuário", placeholder="seu.login")
            password = st.text_input("Senha", type="password", placeholder="••••••••")
            submit   = st.form_submit_button("Entrar →", use_container_width=True)

            if submit:
                user = next((u for u in st.session_state.users
                             if u["login"] == login and u["password"] == password), None)
                if user:
                    st.session_state.logged_in    = True
                    st.session_state.current_user = user
                    st.rerun()
                else:
                    st.error("Usuário ou senha incorretos.")

        st.markdown("""
        <div style="text-align:center;margin-top:16px;color:rgba(255,255,255,0.5);font-size:12px">
            Admin: <strong style="color:white">admin</strong> / <strong style="color:white">admin123</strong>
            &nbsp;·&nbsp; Reps: senha padrão <strong style="color:white">123456</strong>
        </div>
        """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════

def render_header(subtitle):
    user = st.session_state.current_user
    col1, col2 = st.columns([3, 1])
    with col1:
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:12px;padding:8px 0 20px">
            <div style="width:42px;height:42px;border-radius:13px;background:#F47920;display:flex;
                align-items:center;justify-content:center;font-size:20px;
                box-shadow:0 4px 16px rgba(244,121,32,0.4)">📦</div>
            <div>
                <div style="font-weight:800;font-size:17px;letter-spacing:-0.01em">Portal do Representante</div>
                <div style="color:rgba(255,255,255,0.6);font-size:12px">{subtitle}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div style="text-align:right;padding:8px 0 4px">
            <div style="font-weight:700;font-size:13px">{user['name'][:30]}</div>
            <div style="color:rgba(255,255,255,0.5);font-size:11px">{user.get('segmento','')}</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Sair", key="logout_btn"):
            st.session_state.logged_in    = False
            st.session_state.current_user = None
            st.rerun()
    st.markdown("<hr style='border:none;border-top:1px solid rgba(255,255,255,0.1);margin:0 0 24px'>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# HTML TABLE HELPER — substitui st.dataframe com visual bonito
# ═══════════════════════════════════════════════════════════════

def html_table(headers, rows, max_rows=300):
    """Renders a styled HTML table matching the app theme."""
    if not rows:
        st.markdown(
            "<div style='color:rgba(255,255,255,0.5);text-align:center;padding:40px 0;font-size:14px'>Sem registros</div>",
            unsafe_allow_html=True,
        )
        return

    th_style = (
        "padding:10px 14px;font-size:10px;font-weight:700;text-transform:uppercase;"
        "letter-spacing:0.09em;color:rgba(255,255,255,0.55);border-bottom:1px solid rgba(255,255,255,0.12);"
        "white-space:nowrap;background:#163e50;text-align:left;"
    )
    td_style = (
        "padding:10px 14px;font-size:12px;border-bottom:1px solid rgba(255,255,255,0.05);"
        "white-space:nowrap;color:white;vertical-align:middle;"
    )
    td_first = td_style + "color:#F47920;font-weight:600;"

    header_html = "".join(f"<th style='{th_style}'>{h}</th>" for h in headers)

    rows_html = ""
    for row in rows[:max_rows]:
        cells = ""
        for j, cell in enumerate(row):
            style = td_first if j == 0 else td_style
            cells += f"<td style='{style}'>{cell}</td>"
        rows_html += (
            f"<tr style='transition:background 0.15s' "
            f"onmouseover=\"this.style.background='rgba(255,255,255,0.04)'\" "
            f"onmouseout=\"this.style.background='transparent'\">{cells}</tr>"
        )

    table_html = f"""
    <div style="overflow-x:auto;border-radius:14px;border:1px solid rgba(255,255,255,0.1);
        background:#1a5060;margin-top:4px">
      <table style="width:100%;border-collapse:collapse;">
        <thead><tr>{header_html}</tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
      {f'<div style="padding:10px 14px;font-size:11px;color:rgba(255,255,255,0.4)">Mostrando {max_rows} de {len(rows)} registros</div>' if len(rows) > max_rows else ''}
    </div>
    """
    st.markdown(table_html, unsafe_allow_html=True)


def badge_html(text, color):
    return (
        f"<span style='background:{color}22;color:{color};"
        f"border:1px solid {color}55;border-radius:20px;"
        f"padding:2px 9px;font-size:10px;font-weight:700;"
        f"letter-spacing:0.04em;text-transform:uppercase'>{text}</span>"
    )


# ═══════════════════════════════════════════════════════════════
# PRODUTO SEARCH
# ═══════════════════════════════════════════════════════════════

def render_produto_search():
    estoque = st.session_state.estoque
    df = pd.DataFrame(estoque)

    section_header("🔍 Busca de Produtos", f"{len(df)} produtos com saldo disponível")

    mode = st.radio("Modo de busca", ["Por SKU / Nome", "Por Configuração"], horizontal=True)

    if mode == "Por SKU / Nome":
        query = st.text_input("SKU, Código ou Descrição", placeholder="Ex: PA.43.21... ou FITA CREPE...")
        if query:
            q = query.lower()
            mask = (df["sku"].str.lower().str.contains(q, na=False) |
                    df["produto"].str.lower().str.contains(q, na=False) |
                    df["desc_produto"].str.lower().str.contains(q, na=False))
            result = df[mask]
        else:
            result = df
    else:
        c1, c2, c3, c4 = st.columns(4)
        familias = ["Todas"] + sorted(df["familia"].dropna().unique().tolist())
        filiais  = ["Todas"] + sorted(df["filial"].dropna().unique().tolist())
        with c1: fam = st.selectbox("Família", familias)
        with c2: fil = st.selectbox("Filial", filiais)
        with c3: larg = st.text_input("Largura (mm)", placeholder="Ex: 25")
        with c4: comp = st.text_input("Comprimento (m)", placeholder="Ex: 50")

        result = df.copy()
        if fam != "Todas": result = result[result["familia"] == fam]
        if fil != "Todas": result = result[result["filial"]  == fil]
        if larg:
            try: result = result[result["largura"] == float(larg)]
            except: pass
        if comp:
            try: result = result[result["comprimento"] == float(comp)]
            except: pass

    st.markdown(f"**{len(result)}** produto{'s' if len(result) != 1 else ''} encontrado{'s' if len(result) != 1 else ''}")

    if not result.empty:
        rows_out = []
        for _, r in result.head(300).iterrows():
            rows_out.append([
                badge_html(r["sku"], "#F47920"),
                r["produto"],
                badge_html(r["familia"], "#4db8d4") if r["familia"] else "—",
                r["filial"],
                f"{r['largura']:.0f} mm" if r["largura"] > 0 else "—",
                f"{r['comprimento']:.0f} m" if r["comprimento"] > 0 else "—",
                f"<span style='color:#8DC63F;font-weight:700'>{r['m2_disponivel']:,.2f} m²</span>",
                f"{r['rolos_disponivel']:,.0f}",
            ])
        html_table(
            ["SKU","Produto","Família","Filial","Largura","Comp.","M² Disp.","Rolos"],
            rows_out, max_rows=300
        )


# ═══════════════════════════════════════════════════════════════
# PERFORMANCE REP
# ═══════════════════════════════════════════════════════════════

def render_performance_rep(vendedor):
    pedidos = st.session_state.pedidos
    my = [p for p in pedidos if p["vendedor"] == vendedor]
    df = pd.DataFrame(my) if my else pd.DataFrame()

    section_header("📈 Minha Performance", "Visão consolidada da sua carteira")

    if df.empty:
        st.info("Nenhum pedido encontrado para este representante.")
        return

    cart_brl = df["valor_vendido"].sum()
    fat_brl  = df["valor_atendido"].sum()
    sal_brl  = df["valor_a_entregar"].sum()
    cart_m2  = df["m2_vendido"].sum()
    fat_m2   = df["m2_atendido"].sum()
    sal_m2   = df["m2_saldo"].sum()
    pct      = (fat_brl / cart_brl * 100) if cart_brl > 0 else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("🛒 Carteira Total",   BRL(cart_brl), M2(cart_m2))
    c2.metric("✅ Atendido",         BRL(fat_brl),  M2(fat_m2))
    c3.metric("⏳ Saldo a Faturar",  BRL(sal_brl),  M2(sal_m2))
    c4.metric("📋 Itens em Carteira", len(df))

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
    prog_bar(pct, BRL(fat_brl), BRL(cart_brl))

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("#### 🏆 Top Produtos Atendidos")
        fat_df = df[df["valor_atendido"] > 0].groupby("produto").agg(
            Valor=("valor_atendido","sum"), M2=("m2_atendido","sum")
        ).sort_values("Valor", ascending=False).head(5)
        if fat_df.empty:
            st.info("Sem atendimentos registrados.")
        else:
            for i, (prod, row) in enumerate(fat_df.iterrows()):
                pct_bar = row["Valor"] / fat_df["Valor"].iloc[0] * 100
                st.markdown(f"""
                <div style="margin-bottom:12px">
                    <div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:4px">
                        <span><span style="color:#F47920;font-weight:700">#{i+1}</span> {prod[:40]}</span>
                        <span style="color:#8DC63F;font-weight:700">{BRL(row['Valor'])}</span>
                    </div>
                    <div style="background:rgba(0,0,0,0.3);border-radius:99px;height:7px">
                        <div style="height:100%;width:{pct_bar}%;background:linear-gradient(90deg,#F47920,#f5c842);border-radius:99px"></div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

    with col_b:
        st.markdown("#### 📊 Status dos Pedidos")
        status_df = df.groupby("status").size().reset_index(name="Qtd").sort_values("Qtd", ascending=False)
        for _, row in status_df.iterrows():
            color = STATUS_COLOR.get(row["status"], "rgba(255,255,255,0.4)")
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;align-items:center;
                padding:8px 0;border-bottom:1px solid rgba(255,255,255,0.05);font-size:13px">
                <span style="color:{color}">{row['status'] or 'Sem Status'}</span>
                <span style="background:{color}22;color:{color};border:1px solid {color}55;
                    border-radius:20px;padding:2px 10px;font-size:11px;font-weight:700">{row['Qtd']}</span>
            </div>
            """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# PERFORMANCE ADMIN
# ═══════════════════════════════════════════════════════════════

def render_performance_admin():
    users   = st.session_state.users
    pedidos = st.session_state.pedidos
    reps    = [u for u in users if u["role"] == "rep"]
    medals  = ["🥇", "🥈", "🥉"]

    section_header("📈 Performance Geral", "Ranking de todos os representantes com pedidos ativos")

    sort_opt = st.radio("Ordenar por", ["Carteira R$", "Atendido R$", "Saldo R$", "% Atend.", "Nº Itens"], horizontal=True)
    sort_map = {"Carteira R$":"cart_brl","Atendido R$":"fat_brl","Saldo R$":"sal_brl","% Atend.":"pct","Nº Itens":"itens"}
    sort_key = sort_map[sort_opt]

    stats = []
    for u in reps:
        my = [p for p in pedidos if p["vendedor"] == u["vendedor"]]
        if not my: continue
        cart_brl = sum(p["valor_vendido"]    for p in my)
        fat_brl  = sum(p["valor_atendido"]   for p in my)
        sal_brl  = sum(p["valor_a_entregar"] for p in my)
        cart_m2  = sum(p["m2_vendido"]       for p in my)
        fat_m2   = sum(p["m2_atendido"]      for p in my)
        pct      = (fat_brl / cart_brl * 100) if cart_brl > 0 else 0
        stats.append({**u, "cart_brl":cart_brl,"fat_brl":fat_brl,"sal_brl":sal_brl,
                      "cart_m2":cart_m2,"fat_m2":fat_m2,"pct":pct,"itens":len(my)})

    stats.sort(key=lambda x: x[sort_key], reverse=True)

    tCart = sum(r["cart_brl"] for r in stats)
    tFat  = sum(r["fat_brl"]  for r in stats)
    tSal  = sum(r["sal_brl"]  for r in stats)

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("🛒 Carteira Total",   BRL(tCart))
    c2.metric("✅ Total Atendido",   BRL(tFat))
    c3.metric("⏳ Saldo a Faturar",  BRL(tSal))
    c4.metric("👥 Representantes",   len(stats), "com pedidos ativos")

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    max_cart = stats[0]["cart_brl"] if stats else 1

    for i, r in enumerate(stats):
        bar_w   = (r["cart_brl"] / max_cart * 100) if max_cart > 0 else 0
        bar_col = "#8DC63F" if i == 0 else "#F47920"
        pct_col = "#8DC63F" if r["pct"]>=70 else "#f5c842" if r["pct"]>=40 else "#e05555"
        medal   = medals[i] if i < 3 else f"#{i+1}"

        st.markdown(f"""
        <div style="background:#1e6070;border-radius:18px;border:1px solid rgba(255,255,255,0.1);
            padding:16px 22px;margin-bottom:10px">
            <div style="display:flex;flex-wrap:wrap;gap:16px;align-items:center">
                <div style="display:flex;align-items:center;gap:12px;flex:1;min-width:200px">
                    <div style="font-size:26px;min-width:32px;text-align:center">{medal}</div>
                    <div>
                        <div style="font-weight:700;font-size:14px">{r['name'][:35]}</div>
                        <div style="color:rgba(255,255,255,0.5);font-size:11px;margin-top:2px">{r['itens']} itens em carteira</div>
                    </div>
                </div>
                <div style="display:flex;gap:24px;flex-wrap:wrap">
                    <div><div style="color:rgba(255,255,255,0.5);font-size:9px;font-weight:700;text-transform:uppercase">Carteira</div>
                         <div style="color:#F47920;font-weight:800;font-size:15px">{BRL(r['cart_brl'])}</div></div>
                    <div><div style="color:rgba(255,255,255,0.5);font-size:9px;font-weight:700;text-transform:uppercase">Atendido</div>
                         <div style="color:#8DC63F;font-weight:800;font-size:15px">{BRL(r['fat_brl'])}</div></div>
                    <div><div style="color:rgba(255,255,255,0.5);font-size:9px;font-weight:700;text-transform:uppercase">Saldo</div>
                         <div style="color:#f5c842;font-weight:800;font-size:15px">{BRL(r['sal_brl'])}</div></div>
                    <div><div style="color:rgba(255,255,255,0.5);font-size:9px;font-weight:700;text-transform:uppercase">% Atend.</div>
                         <div style="color:{pct_col};font-weight:800;font-size:15px">{r['pct']:.1f}%</div></div>
                </div>
            </div>
            <div style="margin-top:12px;background:rgba(0,0,0,0.3);border-radius:99px;height:7px">
                <div style="height:100%;width:{bar_w}%;background:linear-gradient(90deg,{bar_col}88,{bar_col});border-radius:99px"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# REP DASHBOARD
# ═══════════════════════════════════════════════════════════════

def render_rep_dashboard():
    user    = st.session_state.current_user
    pedidos = st.session_state.pedidos
    estoque = st.session_state.estoque

    render_header("Meu Painel")

    my  = [p for p in pedidos if p["vendedor"] == user["vendedor"]]
    df  = pd.DataFrame(my) if my else pd.DataFrame()
    est = pd.DataFrame(estoque)

    # KPIs topo
    cart_brl = df["valor_vendido"].sum()   if not df.empty else 0
    fat_brl  = df["valor_atendido"].sum()  if not df.empty else 0
    sal_brl  = df["valor_a_entregar"].sum()if not df.empty else 0
    cart_m2  = df["m2_vendido"].sum()      if not df.empty else 0
    sal_m2   = df["m2_saldo"].sum()        if not df.empty else 0
    est_m2   = est["m2_disponivel"].sum()

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("🏭 Estoque Total",    M2(est_m2))
    c2.metric("🛒 Carteira Total",   BRL(cart_brl), M2(cart_m2))
    c3.metric("✅ Atendido",         BRL(fat_brl))
    c4.metric("⏳ Saldo a Faturar",  BRL(sal_brl),  M2(sal_m2))

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs(["🛒 Minha Carteira", "🏭 Estoque", "🔍 Produtos", "📈 Performance"])

    with tab1:
        if df.empty:
            st.info("Nenhum pedido em carteira.")
        else:
            STATUS_C = {"Aguardando Faturamento":"#f5c842","Aguardando Faturamento - Parcial":"#f5c842",
                        "Pedido de Venda em Aberto":"#4db8d4","Aguardando Separação":"#F47920",
                        "Financeiro Rejeitado":"#e05555","Aguardando Liberação Estoque":"#b07aff",
                        "Aguardando Liberação Comercial":"#b07aff"}
            rows_out = []
            for _, p in df.iterrows():
                sc = STATUS_C.get(p["status"], "rgba(255,255,255,0.4)")
                rows_out.append([
                    p["pedido"],
                    badge_html(p["sku"], "#F47920"),
                    p["cliente"][:28],
                    p["produto"][:28],
                    badge_html(p["status"], sc),
                    p["uf"],
                    f"{p['m2_vendido']:,.2f}",
                    f"{p['m2_atendido']:,.2f}",
                    f"{p['m2_saldo']:,.2f}",
                    f"<span style='color:#F47920'>{BRL(p['valor_vendido'])}</span>",
                    f"<span style='color:#8DC63F'>{BRL(p['valor_atendido'])}</span>",
                    f"<span style='color:#f5c842'>{BRL(p['valor_a_entregar'])}</span>",
                    p["emissao"], p["entrega"],
                ])
            html_table(["Pedido","SKU","Cliente","Produto","Status","UF",
                        "m² Vend.","m² Aten.","m² Saldo",
                        "R$ Vendido","R$ Atend.","R$ Saldo","Emissão","Entrega"], rows_out)

    with tab2:
        rows_out = []
        for e in estoque[:300]:
            rows_out.append([
                badge_html(e["sku"], "#F47920"),
                e["produto"],
                badge_html(e["familia"], "#4db8d4") if e["familia"] else "—",
                e["filial"],
                f"{e['largura']:.0f} mm" if e["largura"] > 0 else "—",
                f"{e['comprimento']:.0f} m" if e["comprimento"] > 0 else "—",
                f"<span style='color:#8DC63F;font-weight:700'>{e['m2_disponivel']:,.2f} m²</span>",
                f"{e['rolos_disponivel']:,.0f}",
            ])
        html_table(["SKU","Produto","Família","Filial","Largura","Comp.","M² Disp.","Rolos"], rows_out, max_rows=300)

    with tab3:
        render_produto_search()

    with tab4:
        render_performance_rep(user["vendedor"])


# ═══════════════════════════════════════════════════════════════
# ADMIN DASHBOARD
# ═══════════════════════════════════════════════════════════════

def render_admin_dashboard():
    render_header("Painel Administrativo")

    pedidos = st.session_state.pedidos
    estoque = st.session_state.estoque
    users   = st.session_state.users
    reps    = [u for u in users if u["role"] == "rep"]

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📊 Visão Geral", "📈 Performance", "🛒 Carteira Geral", "🔍 Produtos", "👥 Representantes"
    ])

    # ── VISÃO GERAL ──────────────────────────────────────────────
    with tab1:
        section_header("📊 Visão Geral")

        df  = pd.DataFrame(pedidos)
        est = pd.DataFrame(estoque)

        tCart = df["valor_vendido"].sum()
        tFat  = df["valor_atendido"].sum()
        tSal  = df["valor_a_entregar"].sum()
        tM2   = est["m2_disponivel"].sum()

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("🛒 Carteira Total",  BRL(tCart))
        c2.metric("✅ Total Atendido",  BRL(tFat))
        c3.metric("⏳ Saldo a Faturar", BRL(tSal))
        c4.metric("🏭 Estoque M²",      M2(tM2))
        c5.metric("👥 Representantes",  len(reps))
        c6.metric("📋 Itens Carteira",  len(df))

        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
        st.markdown("#### Resumo por Representante")

        rows_out = []
        for u in reps:
            my = df[df["vendedor"] == u["vendedor"]] if not df.empty else pd.DataFrame()
            cv = my["valor_vendido"].sum()    if not my.empty else 0
            fv = my["valor_atendido"].sum()   if not my.empty else 0
            sv = my["valor_a_entregar"].sum() if not my.empty else 0
            cm = my["m2_vendido"].sum()       if not my.empty else 0
            sm = my["m2_saldo"].sum()         if not my.empty else 0
            rows_out.append([
                u["name"][:35],
                len(my),
                f"<span style='color:#F47920'>{BRL(cv)}</span>",
                f"<span style='color:#8DC63F'>{BRL(fv)}</span>",
                f"<span style='color:#f5c842'>{BRL(sv)}</span>",
                f"{cm:,.2f} m²",
                f"{sm:,.2f} m²",
            ])
        html_table(["Representante","Itens","Carteira R$","Atendido R$","Saldo R$","m² Cart.","m² Saldo"], rows_out)

    # ── PERFORMANCE ──────────────────────────────────────────────
    with tab2:
        render_performance_admin()

    with tab3:
        section_header("🛒 Carteira Geral", f"{len(pedidos)} itens no total")
        if pedidos:
            STATUS_C = {"Aguardando Faturamento":"#f5c842","Aguardando Faturamento - Parcial":"#f5c842",
                        "Pedido de Venda em Aberto":"#4db8d4","Aguardando Separação":"#F47920",
                        "Financeiro Rejeitado":"#e05555","Aguardando Liberação Estoque":"#b07aff",
                        "Aguardando Liberação Comercial":"#b07aff"}
            rows_out = []
            for p in pedidos:
                sc = STATUS_C.get(p["status"], "rgba(255,255,255,0.4)")
                rows_out.append([
                    p["pedido"],
                    badge_html(p["sku"], "#F47920"),
                    p["cliente"][:22],
                    p["produto"][:22],
                    p["vendedor"][:22],
                    badge_html(p["status"], sc),
                    p["uf"],
                    f"{p['m2_vendido']:,.2f}",
                    f"{p['m2_saldo']:,.2f}",
                    f"<span style='color:#F47920'>{BRL(p['valor_vendido'])}</span>",
                    f"<span style='color:#f5c842'>{BRL(p['valor_a_entregar'])}</span>",
                    p["emissao"],
                ])
            html_table(["Pedido","SKU","Cliente","Produto","Vendedor","Status","UF",
                        "m² Vend.","m² Saldo","R$ Vendido","R$ Saldo","Emissão"], rows_out, max_rows=300)

    # ── PRODUTOS ─────────────────────────────────────────────────
    with tab4:
        render_produto_search()

    # ── REPRESENTANTES ───────────────────────────────────────────
    with tab5:
        section_header("👥 Gestão de Usuários", f"{len(users) - 1} usuário(s) cadastrado(s)")

        # ── Upload planilha ──────────────────────────────────────
        with st.expander("📁 Atualizar dados — Upload de Planilha Excel"):
            st.markdown("Suba a planilha `.xlsx` com abas **Estoque** e **Pedidos** para atualizar os dados.")
            uploaded = st.file_uploader("Escolha o arquivo .xlsx", type=["xlsx","xls"])
            if uploaded:
                with st.spinner("Processando planilha..."):
                    try:
                        novo_est, novo_ped = load_from_excel(uploaded.read())
                        st.session_state.estoque = novo_est
                        st.session_state.pedidos = novo_ped
                        st.success(f"✅ {len(novo_est)} produtos e {len(novo_ped)} pedidos carregados!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao processar: {e}")

        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

        # ── Formulário novo usuário ──────────────────────────────
        with st.expander("➕ Cadastrar novo usuário"):
            with st.form("new_user_form", clear_on_submit=True):
                st.markdown("#### Novo usuário")
                c1, c2 = st.columns(2)
                with c1:
                    nome     = st.text_input("Nome completo *")
                    vendedor = st.text_input("Vendedor (igual à planilha) *",
                                             help="Deve ser idêntico ao nome na coluna Vendedor da planilha de pedidos")
                    segmento = st.text_input("Segmento", placeholder="Ex: DEFIR")
                with c2:
                    login    = st.text_input("Login *", placeholder="Ex: joao.silva")
                    password = st.text_input("Senha *", type="password")
                    role     = st.selectbox("Perfil", ["rep", "admin"],
                                            format_func=lambda x: "👤 Representante" if x == "rep" else "🔑 Administrador")
                salvar = st.form_submit_button("✅ Salvar usuário", use_container_width=True)
                if salvar:
                    if not all([nome, vendedor, login, password]):
                        st.error("⚠️ Preencha todos os campos obrigatórios (*)")
                    elif any(u["login"] == login for u in st.session_state.users):
                        st.error("⚠️ Esse login já existe. Escolha outro.")
                    else:
                        st.session_state.users.append({
                            "id":       max(u["id"] for u in st.session_state.users) + 1,
                            "name":     nome,
                            "vendedor": vendedor,
                            "segmento": segmento,
                            "login":    login,
                            "password": password,
                            "role":     role,
                        })
                        st.success(f"✅ Usuário **{nome}** cadastrado com sucesso!")
                        st.rerun()

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Edição inline de senha ───────────────────────────────
        if "editing_pwd" not in st.session_state:
            st.session_state.editing_pwd = None

        # ── Lista de usuários ────────────────────────────────────
        outros = [u for u in st.session_state.users if u["id"] != 1]

        # Header da lista
        st.markdown(f"""
        <div style="display:grid;grid-template-columns:2fr 1.5fr 1fr 1fr auto auto auto;
            gap:8px;padding:8px 14px;
            background:#163e50;border-radius:10px 10px 0 0;
            border:1px solid rgba(255,255,255,0.1);
            font-size:10px;font-weight:700;text-transform:uppercase;
            letter-spacing:0.08em;color:rgba(255,255,255,0.55);margin-bottom:0">
            <span>Nome</span><span>Login</span><span>Vendedor</span>
            <span>Perfil</span><span style="text-align:center">Senha</span>
            <span style="text-align:center">Perfil</span>
            <span style="text-align:center">Remover</span>
        </div>
        """, unsafe_allow_html=True)

        for u in outros:
            role_color = "#F47920" if u["role"] == "admin" else "#8DC63F"
            role_label = "🔑 Admin" if u["role"] == "admin" else "👤 Rep"
            toggle_label = "↓ Rep" if u["role"] == "admin" else "↑ Admin"

            c1, c2, c3, c4, c5, c6, c7 = st.columns([2, 1.5, 1, 1, 0.8, 0.8, 0.8])

            c1.markdown(f"<div style='font-weight:700;font-size:13px;padding-top:6px'>{u['name'][:30]}</div>",
                        unsafe_allow_html=True)
            c2.markdown(f"<div style='color:#4db8d4;font-size:12px;padding-top:6px'>{u['login']}</div>",
                        unsafe_allow_html=True)
            c3.markdown(f"<div style='color:rgba(255,255,255,0.5);font-size:11px;padding-top:6px'>{u['vendedor'][:18]}</div>",
                        unsafe_allow_html=True)
            c4.markdown(
                f"<span style='background:{role_color}22;color:{role_color};"
                f"border:1px solid {role_color}55;border-radius:20px;"
                f"padding:3px 10px;font-size:11px;font-weight:700'>{role_label}</span>",
                unsafe_allow_html=True,
            )

            # Trocar senha
            with c5:
                if st.button("🔑", key=f"pwd_{u['id']}", help="Alterar senha"):
                    st.session_state.editing_pwd = u["id"] if st.session_state.editing_pwd != u["id"] else None
                    st.rerun()

            # Promover / Rebaixar
            with c6:
                if st.button(toggle_label, key=f"role_{u['id']}", help="Alterar perfil"):
                    for usr in st.session_state.users:
                        if usr["id"] == u["id"]:
                            usr["role"] = "rep" if usr["role"] == "admin" else "admin"
                    st.rerun()

            # Remover
            with c7:
                if st.button("✕", key=f"del_{u['id']}", help="Remover usuário"):
                    st.session_state.users = [x for x in st.session_state.users if x["id"] != u["id"]]
                    if st.session_state.editing_pwd == u["id"]:
                        st.session_state.editing_pwd = None
                    st.rerun()

            # Formulário inline de troca de senha
            if st.session_state.editing_pwd == u["id"]:
                with st.form(key=f"form_pwd_{u['id']}"):
                    st.markdown(f"<div style='color:rgba(255,255,255,0.6);font-size:12px;margin-bottom:4px'>Nova senha para <strong>{u['name']}</strong></div>",
                                unsafe_allow_html=True)
                    nova_senha = st.text_input("Nova senha", type="password", key=f"ns_{u['id']}")
                    conf_senha = st.text_input("Confirmar senha", type="password", key=f"cs_{u['id']}")
                    if st.form_submit_button("Salvar nova senha"):
                        if not nova_senha:
                            st.error("Digite a nova senha.")
                        elif nova_senha != conf_senha:
                            st.error("As senhas não coincidem.")
                        elif len(nova_senha) < 4:
                            st.error("Senha muito curta (mínimo 4 caracteres).")
                        else:
                            for usr in st.session_state.users:
                                if usr["id"] == u["id"]:
                                    usr["password"] = nova_senha
                            st.session_state.editing_pwd = None
                            st.success("✅ Senha atualizada!")
                            st.rerun()

            st.markdown("<hr style='border:none;border-top:1px solid rgba(255,255,255,0.05);margin:4px 0'>",
                        unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    init_state()

    if not st.session_state.logged_in:
        render_login()
        return

    user = st.session_state.current_user
    # Atualiza dados do usuário se mudou (ex: role alterado)
    updated = next((u for u in st.session_state.users if u["id"] == user["id"]), None)
    if updated:
        st.session_state.current_user = updated

    if st.session_state.current_user["role"] == "admin":
        render_admin_dashboard()
    else:
        render_rep_dashboard()


if __name__ == "__main__":
    main()
